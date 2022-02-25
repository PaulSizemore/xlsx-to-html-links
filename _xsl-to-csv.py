# dev-xsl-to-csv.py
# extracts all the sheets with a URL
#
# Issues: select all, select none, open selected
# deselect certain domains
import csv
import os
import pandas as pd
import re
import openpyxl


file_dir = '/Users/paulb.sizemore/OneDrive/_github/data_csv-to-html-links/'

files = [  f for f in os.listdir(file_dir)  if f.endswith(".xlsx")]

#########################################
# Loop the Excel Files
for i in files:
    siteName = i.replace(".xlsx","")
    siteDirectory = file_dir   # file_dir + siteName + '/'

    if not os.path.exists(siteDirectory):
        os.makedirs(siteDirectory)
    #print('-----------------------------------------------')
    #print(i)
    #print('-----------------------------------------------')
    #file_name = file_dir + i.replace(".xlsx","") + ".csv"
    file_name = file_dir + i
    wb=openpyxl.load_workbook(file_name)
    sheetNames = wb.sheetnames
    # also sheetnames from ws = wb.worksheets[0]

    #########################################
    # Loop the sheet names in the Excel file
    for ws in wb:

        ###########################
        # Overwrite the CSV file
        file_name_out = siteDirectory + i.replace('.xlsx','_')  + ws.title.lower()  + '.csv'

        print(file_name_out)
        rowOutput = ''

        for row in ws.values:
            for value in row:
                cleanValue = str(value)
                cleanValue = cleanValue.replace('"','\\"')
                rowOutput = str(rowOutput) + '"' + cleanValue + '",'

            rowOutput = rowOutput + '""\n'

        with open(file_name_out, "w") as text_file:
            print(rowOutput, file=text_file)
        




